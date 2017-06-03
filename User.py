from Requirement import Requirement
from Course import Course
from ast import literal_eval
from openpyxl import load_workbook
from Authentication import authenticate
from copy import deepcopy, copy
from urllib.request import urlopen

'''
TODO:
1. PORT INTO ANOTHER LANGUAGE BECAUSE PYTHON SUCKS AT MULTITASKING
    a) Multi-thread
    b) Prioritize taken programs

2. Load in all the programs
    a) Hell on earth
'''

PROGRAM_FILE = 'data\data.dat'  #'data\\testprograms.txt'
DEFAULT_PROGRAMS = literal_eval(open(PROGRAM_FILE).readline())
FILE_NAME = 'data\\database.xlsx'
PAGE = 'https://cobalt.qas.im/api/1.0/courses/filter?q=code:%22'
KEY = 'TVwEIjRZP80vhnY8HhM0OzZCMfydh4lA'
WORKBOOK = load_workbook(FILE_NAME)
WORKSHEET = WORKBOOK.active
USERS = {}


def get_user_lines():
    cell = 0
    row = 2

    while cell is not None:
        cell = WORKSHEET.cell(column=1, row=row).value

        USERS[cell] = row
        row += 1

get_user_lines()


class User:
    """
    A unique-user interactive interface
    
    Public Attributes:
    =================
        @param str username: 
            User's username
    
    Private Attributes:
    ==================
        @param bool _logged_in: 
            Whether or not the user successfully logged in
        @param list _taken_courses:
            A list of Course objects that the user has added
        @param list _taken_programs:
            A list of program codes that the user has added
        @param dict _program_requirement_cache:
            A dictionary of Requirement objects for their respective programs
    """
    def __init__(self, username, password):
        if not authenticate(username.lower(), password, USERS, WORKBOOK, WORKSHEET):
            self._logged_in = False
        else:
            self._logged_in = True
            self.username = username.lower()

            self._taken_courses = {}
            self._taken_programs = []
            self._program_requirement_cache = {}
            self._breadths = [0, 0, 0, 0, 0]

            self.initial_courses()
            self.initial_programs()
            self.initial_program_cache()

    def logged_in(self):
        """
        Return if the user is logged in or not

        :return: bool
        """

        return self._logged_in

    def _update_requirements(self):
        """
        Update all program requirement objects

        :return: None
        """

        for requirement in self._program_requirement_cache:
            self._program_requirement_cache[requirement].update

    # ============= ADD METHODS ================== #

    def add_program(self, program_code):
        """
        Add <program_code> to the user's taken programs

        :param str program_code: Program code to add
        :return: None
        """

        if program_code not in self._taken_programs:
            WORKSHEET.cell(column=3, row=USERS[self.username], value=str(self._taken_programs))
            WORKBOOK.save(FILE_NAME)

        else:
            print('Program already added!')

    def add_course(self, course_code, initial=False):
        """
        Add <course_code> to the user's taken courses

        :param str course_code: Course to add
        :param bool initial: If doing the initial startup - don't bother updating all the requirements since they aren't made yet
        :return: None
        """

        if course_code not in self._taken_courses:
            lines = urlopen("https://timetable.iit.artsci.utoronto.ca/api/20179/courses?code=" + course_code).readlines()  # Get course info

            if len(lines) < 1:
                lines = urlopen("https://timetable.iit.artsci.utoronto.ca/api/20169/courses?code=" + course_code).readlines()  # Get course info

            if len(lines) > 1:
                breadths = self._get_breadths(lines[15].decode('utf-8'))

                for breadth in breadths:
                    self._breadths[breadth - 1] += 1

                # Local courses
                self._taken_courses[course_code] = Course(course_code, breadths=breadths)

                # Database courses
                WORKSHEET.cell(column=2, row=USERS[self.username], value=str([[course_code, self._taken_courses[course_code].type, self._taken_courses[course_code].mark] for course_code in self._taken_courses]))
                WORKBOOK.save(FILE_NAME)
                if not initial:
                    self._update_requirements()

            else:
                print('Couldn\'t find course!')

    # ============= REMOVE METHODS ================== #

    def remove_program(self, program_code):
        """
        Remove <program_code> from user's taken programs

        :param str program_code: Program to remove
        :return: None
        """

        try:
            self._taken_programs.remove(program_code)

            WORKSHEET.cell(column=3, row=USERS[self.username], value=str(self._taken_programs))
            WORKBOOK.save(FILE_NAME)

        except KeyError:
            print('Program does not exist!')

    def remove_course(self, course_code):
        """
        Remove <course_code> from user's taken courses

        :param str course_code: Course to remove
        :return: None
        """

        try:
            for breadth in self._taken_courses[course_code].breadths:
                self._breadths[breadth] -= 1

            # Local courses
            self._taken_courses.remove(course_code)

            # Database courses
            WORKSHEET.cell(column=2, row=USERS[self.username], value=str([[course_code, self._taken_courses[course_code].type, self._taken_courses[course_code].mark] for course_code in self._taken_courses]))
            WORKBOOK.save(FILE_NAME)
            self._update_requirements()

        except Exception:
            pass

    # ============= GET METHODS ================== #

    def get_programs(self):
        """
        Return program names

        :return: list of str
        """

        return self._taken_programs

    def get_program(self, program_code):
        """
        Return <program_code>'s program information

        :param str program_code: Program to get information from
        :return: dict of str
        """

        temp = copy(DEFAULT_PROGRAMS[program_code])
        temp.pop('requirements')
        return temp

    def get_program_requirements(self, program_code):
        """
        Return Requirement for <program_code>

        :param str program_code: Program to get requirements for
        :return: Requirement
        """

        return self._program_requirement_cache[program_code]

    def get_total_breadth(self):
        return self._breadths

    def _get_breadths(self, text):
        """
        Return list of int representing the breadths in text

        :param str text: A line of results
        :return: list of int
        """

        start = 0
        breadths = []

        while start != -1:
            start = text.find('(', start + 1)
            if start != -1:
                breadths.append(int(text[start + 1]))

        return breadths

    def get_courses(self):
        """
        Return a list of all courses the user is taking

        :return: list of str
        """
        return self._taken_courses

    def get_course(self, course_code):
        """
        Return <course_code>'s course object

        :param str course_code: Course data is desired of
        :return: Course
        """

        return self._taken_courses[course_code]

    def get_mark(self, course_code):
        """
        Return mark for <course_code>

        :param str course_code: Course desired mark is in
        :return: float
        """

        return self._taken_courses[course_code].mark

    def get_type(self, course_code):
        """
        Return type for <course_code>

        :param str course_code: Course desired type is in
        :return: str
        """

        return self._taken_courses[course_code].type

    def get_easiest(self, limit=5):
        """
        Return a list of easiest to get programs by percentage and courses needed

        :return: dict of list
        """

        values = {'percentage': [], 'to_finish': []}

        for program in self._program_requirement_cache:  # For every program
            shortened = self._program_requirement_cache[program]

            percentage = (shortened.have / shortened.need if shortened.need != 0 else 1) * 100  # Get percentage completed
            to_finish = shortened.need - shortened.have  # Get amount of requirements needed to finish

            values['percentage'].append([DEFAULT_PROGRAMS[program]['name'], percentage])
            values['to_finish'].append([DEFAULT_PROGRAMS[program]['name'], to_finish])

        values['percentage'].sort(reverse=True, key=lambda x: x[1])  # Sort from Highest -> Lowest
        values['to_finish'].sort(key=lambda x: x[1])  # Sort from Least -> Most

        values['percentage'] = values['percentage'][:limit]
        values['to_finish'] = values['to_finish'][:limit]

        return values

    def get_cgpa(self):
        """
        Return current grade point average from courses taken

        :return: float
        """

        total = 0
        credits = 0

        for course in self._taken_courses:
            course = self._taken_courses[course]
            if course.type != 'Dropped':
                # If the course wasn't dropped

                mark = course.mark
                if mark > 84:
                    grade_point = 4.0
                elif mark > 79:
                    grade_point = 3.7
                elif mark > 76:
                    grade_point = 3.3
                elif mark > 72:
                    grade_point = 3.0
                elif mark > 69:
                    grade_point = 2.7
                elif mark > 66:
                    grade_point = 2.3
                elif mark > 62:
                    grade_point = 2.0
                elif mark > 59:
                    grade_point = 1.7
                elif mark > 56:
                    grade_point = 1.3
                elif mark > 52:
                    grade_point = 1.0
                elif mark > 49:
                    grade_point = 0.7
                else:
                    grade_point = 0

                if course.weight == 1:
                    total += 1 * grade_point
                    credits += 1
                else:
                    total += .5 * grade_point
                    credits += .5

        if total != 0:
            return total / credits
        return 0.0

    # ============= SET METHODS ================== #

    def set_mark(self, course_code, mark):
        """
        Set the mark of <course_code> to <mark>

        :param str course_code: Course to change
        :param float mark: Mark of course
        :return: None
        """

        if len(course_code) == 8 and isinstance(course_code, str) and course_code in self._taken_courses:

            self._taken_courses[course_code].mark = mark

            WORKSHEET.cell(column=2, row=USERS[self.username], value=str([[course_code, self._taken_courses[course_code].type, self._taken_courses[course_code].mark] for course_code in self._taken_courses]))
            WORKBOOK.save(FILE_NAME)

            self._taken_courses[course_code]._update_passed()
            self._update_requirements()

        else:
            print('Course does not exist!')

    def set_type(self, course_code, type_):
        """
        Set the type of <course_code> to <type>

        :param str course_code: Course to change
        :param str type: Type of course
        :return: None
        """

        if len(course_code) == 8 and isinstance(course_code, str) and course_code in self._taken_courses:

            self._taken_courses[course_code].type = type_

            WORKSHEET.cell(column=2, row=USERS[self.username], value=str([[course_code, self._taken_courses[course_code].type, self._taken_courses[course_code].mark] for course_code in self._taken_courses]))
            WORKBOOK.save(FILE_NAME)

            self._taken_courses[course_code]._update_passed()
            self._update_requirements()

        else:
            print('Course does not exist!')

    # ============= INITIAL OBJECT METHODS ================== #

    def initial_courses(self):
        """
        Add all pre-existing user programs

        :return: None
        """
        for course_info in literal_eval(WORKSHEET.cell(column=2, row=USERS[self.username]).value):
            self.add_course(course_info[0], initial=True)
            self.set_type(course_info[0], course_info[1])
            self.set_mark(course_info[0], course_info[2])

    def initial_programs(self):
        """
        Add all pre-existing user programs

        :return: None
        """

        for program_code in literal_eval(WORKSHEET.cell(column=3, row=USERS[self.username]).value):
            self.add_program(program_code)

    def initial_program_cache(self):
        """
        Convert all the pre-existing user data into requirement objects

        :return: None
        """

        for program in DEFAULT_PROGRAMS:
            self._program_requirement_cache[program] = self.convert_to_requirement(deepcopy(DEFAULT_PROGRAMS[program]['requirements']))

        self._update_requirements()

    def convert_to_requirement(self, requirements, exclusions=None, treatall=False, only_used=False, only_unused=False):
        """
        Converts data into an requirement object

        :param list of str requirements: Courses that are needed for the program
        :param list of str|None exclusions: Courses that cannot be used for this part of the requirement
        :param bool treatall: If all courses are being looked through and marked as used
        :param bool only_used: If only courses that have previously been used are being used
        :param bool only_unused: If only courses that have previously not been used are being used
        :return: Requirement
        """

        modifier = requirements[0]
        start_index = 1
        end_index = len(requirements) - 1
        exclusions = set([]) if exclusions is None else exclusions

        if modifier != 'MARK':
            transformed_requirements = []

            if modifier[0] == 'X':
                only_used = True
                only_unused = False
            elif modifier[-1] == 'X':
                only_used = False
                only_unused = True

            max = requirements[-1]
            if isinstance(requirements[-2], int) or isinstance(requirements[-2], float):
                min = requirements[-2]
                end_index -= 1
            else:
                min = None

            if isinstance(requirements[start_index], tuple):
                for item in requirements[start_index]:
                    exclusions.add(item)
                start_index += 1

            if treatall:
                pass
            elif requirements[start_index] == 'TREATALL':
                start_index += 1
                treatall = True
            else:
                treatall = False

            for item in range(start_index, end_index):
                item = requirements[item]

                if isinstance(item, list):
                    transformed_requirements.append(self.convert_to_requirement(item, exclusions=exclusions, treatall=treatall, only_used=only_used, only_unused=only_unused))

                else:
                    transformed_requirements.append(item)

            return Requirement(modifier, min, max, transformed_requirements, exclusions, treatall, only_used, only_unused, self._taken_courses, self._breadths)

        else:
            return Requirement(modifier, None, requirements[-1], requirements[1], exclusions, treatall, only_used, only_unused, self._taken_courses, self._breadths)


if __name__ == '__main__':
    import time

    start = time.time()
    while True:
        start = time.time()
        usr = User(input('Username: ').lower(), input('Password: '))

        if usr.logged_in():
            print('Startup Time:', time.time() - start, 'seconds')
            print('Logged in')
            break
        else:
            print('Incorrect login!\n')
            del usr

    usr.set_mark('CSC108H1', 85)
    usr.set_mark('CSC148H1', 94)
    usr.set_mark('CSC165H1', 80)
    usr.set_mark('MAT135H1', 68)
    usr.set_mark('MAT136H1', 62)
    usr.set_mark('MAT223H1', 57)
    usr.set_mark('PHL101Y1', 70)
    usr.set_mark('SII199Y1', 84)

    start = time.time()
    usr.add_course('MAT224H1')
    print('Course Add Time:', time.time() - start, 'seconds')
    start = time.time()
    usr.remove_course('MAT224H1')
    print('Course Remove Time:', time.time() - start, 'seconds')
    start = time.time()
    print(usr.get_easiest())
    print('Easiest program Time:', time.time() - start, 'seconds')
