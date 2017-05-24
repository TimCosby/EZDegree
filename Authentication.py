from openpyxl import load_workbook

WORKBOOK = load_workbook('data\\authentication.xlsx')
WORKSHEET = WORKBOOK.active

def authenticate(username, password, users, workbook, worksheet):
    """
    
    @param str username: 
    @param str password: 
    @param dict of int users:
    @param Workbook workbook:
    @param Worksheet worksheet:
    @return: bool
    """

    try:
        return password == WORKSHEET.cell(column=2, row=users[username]).value
    except KeyError:
        if 'y' in input('Create a new user?\n').lower():
            row = max(users.values())

            users[username] = row

            WORKSHEET.cell(column=1, row=row, value=str(username))
            WORKSHEET.cell(column=2, row=row, value=str(password))
            WORKBOOK.save('data\\authentication.xlsx')

            worksheet.cell(column=1, row=row, value=username)
            worksheet.cell(column=2, row=row, value=str({}))
            worksheet.cell(column=3, row=row, value=str([]))
            workbook.save('data\\database.xlsx')

            return True
        else:
            return False
