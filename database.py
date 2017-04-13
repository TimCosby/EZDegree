from openpyxl import load_workbook
from openpyxl import Workbook
import ast
import degree_search
from urllib.request import Request
from urllib.request import urlopen
from urllib.request import HTTPError
from urllib.request import URLError
from Course import *
from copy import deepcopy
from build_tree import build_tree

##### TO DO ######
# 1. Deal with adding to program requirements that have ****s in them (ie. USA4****) DEATH
#
# 2. Add breadth requirements to get easiest

FILE_NAME = 'database.xlsx'
PAGE = 'https://cobalt.qas.im/api/1.0/courses/filter?q=code:%22'
KEY = 'TVwEIjRZP80vhnY8HhM0OzZCMfydh4lA'
WORKBOOK = load_workbook(FILE_NAME)
WORKSHEET = WORKBOOK.active
PROGRAM_TREE = eval(open('data.dat', 'r').readline())


class Login:
    def __init__(self, username, password):
        self.username = username.lower()
        self.password = password

        self._user_row = None

        if self._valid_characters():
            response = self._find_user()
            if response[0]:
                # If valid user

                if WORKSHEET.cell(column=2, row=response[1]).value == self.password:
                    # If password is correct
                    self._user_row = response[1]

            elif 'y' in input('Do you want to create a new user? Y/N').lower():
                self._user_row = response[1]

    @property
    def get_row(self):
        return self._user_row

    def _valid_characters(self):
        if ' ' in self.username:
            print('Invalid username characters!')
            return False

        elif ' ' in self.password:
            print('Invalid password characters!')
            return False

        else:
            return True

    def _find_user(self):
        """
        Return if <username> is registered.

        @return: bool
        """

        cell = 0
        row = 1

        while cell is not None:
            # While a name is in the cell
            cell = WORKSHEET.cell(column=1, row=row).value

            if cell == self.username:
                # If the cell is the username
                return True, row

            else:
                row += 1

        # If the names did not match the username
        return False, row - 1


class User:
    """
    Private Attributes:
    ==================
        @param Workbook _wb: 
                    Database workbook
       
        @param Worksheet _ws:
                    Database's worksheet
                    
        @param bool _logged_in:
                    Return if user is logged in
                    
        @param int _user_row:
                    Row that the user is located in worksheet
                    
        @param dict _course_list:
                    List of all courses user has taken
                    
    Public Attributes:
    =================
        @param str username:
                    User's name
    """

    def __init__(self, username, password):
        """
        User's Profile
        
        @param str username: 
        @param str password: 
        """

        login_response = Login(username, password).get_row

        if login_response is not None:
            print('Successfully logged in!')

            self._logged_in = True
            self._user_row = login_response
            self._course_list = {}

            # Get user programs
            self._program_list = ast.literal_eval(WORKSHEET.cell(column=4, row=self._user_row).value)

            self._program_course_cache = {}
            self._program_tree = deepcopy(PROGRAM_TREE)  # Get default program tree
            print(self._program_tree)
            self._user_program_tree = build_tree(self._program_tree, self._program_course_cache)  # Program the degree requirements while caching like courses as the same object
            print(self._program_tree)

            # Put together user's profile of courses
            temp = ast.literal_eval(WORKSHEET.cell(column=3, row=self._user_row).value)
            self._add_pre_courses(temp)

            print('protree',self._user_program_tree)

        else:
            print('Incorrect Login!')
            self._logged_in = False

    """Course Functions"""

    def add_course(self, course_code, lecture_code, ctype, cmark):
        """
        Adds a course to the user's profile
        
        @param str course_code: Course code of the course
        @param str lecture_code: Lecture code
        @param str ctype: Type of course Planned/Taken/etc
        @param float cmark: Mark in the course
        @return: NoneType
        """

        # See if course exists
        course_info = self._get_course_info(course_code[:8])

        # If course does exist
        if course_info is not None and course_info != []:
            # Get rid of useless info
            course_info.pop('id')

            # See if path is in the dictionary
            try:
                if course_code[3:6] in self._course_list[course_code[:3]]:
                    # If code is already in the list - modify
                    self._course_list[course_code[:3]][course_code[3:6]].modify(lecture_code=lecture_code, ctype=ctype, cmark=cmark)
                else:
                    # If course prefix but not code is already in the list, add to list
                    self._course_list[course_code[:3]].update({course_code[3:6]: Course(course_info, lecture_code=lecture_code, ctype=ctype, cmark=cmark)})

            except KeyError:
                # If node does not exist
                self._course_list[course_code[:3]] = {course_code[3:6]: Course(course_info, lecture_code=lecture_code, ctype=ctype, cmark=cmark)}

            WORKSHEET.cell(column=3, row=self._user_row, value=str(self._course_list))

            self._update_program_course(self._course_list[course_code[:3]][course_code[3:6]])
            self._update_program_nodes()

            WORKBOOK.save(FILE_NAME)

        else:
            print('Invalid course name!')

    def remove_course(self, course_code):
        """
        Delete a course from the list
        
        @param str course_code: Course code (ie. CSC148) 
        @return: NoneType
        """

        self._update_program_course(self._course_list[course_code[:3]][course_code[3:6]], False)
        del self._course_list[course_code[:3]][course_code[3:6]]
        self._update_program_nodes()

        # Update database
        WORKSHEET.cell(column=3, row=self._user_row, value=str(self._course_list))
        WORKBOOK.save(FILE_NAME)

    def get_loggedin(self):
        """
        Return if user is logged in

        @return: bool 
        """

        return self._logged_in

    def get_courses(self):
        """
        Return all Courses
        
        @return: list of Course 
        """
        rlist = []
        keys = self._course_list.keys()

        for key in keys:
            for item in self._course_list[key]:
                rlist.append(self._course_list[key][item])

        return rlist

    def get_breadths(self):
        """
        Return breadth totals of planned and taken
        
        @return: dict of int
        """
        breadth_totals = {'Taken': [0, 0, 0, 0, 0], 'Planned': [0, 0, 0, 0, 0]}
        courses = self.get_courses()

        for course in courses:
            course_type = course.ctype

            for breadth in course.breadths:
                # C for completed
                if course_type == 'p':
                    breadth_totals['Planned'][breadth - 1] += course.cworth
                elif not (course_type == 'c' and course.cmark > 50) and course_type != 'd':
                    breadth_totals['Taken'][breadth - 1] += course.cworth

        return breadth_totals

    def get_course_mark(self, course):
        return self._course_list[course[:3]][course[3:6]].cmark

    def get_cgpa(self):
        total = 0
        credits = 0

        courses = self.get_courses()
        for course in courses:
            if course.ctype != 'd':
                # If the course wasn't dropped

                mark = course.cmark
                if mark > 84:
                    grade_point = 4.0
                elif mark > 79:
                    grade_point = 3.7
                elif mark > 76:
                    grade_point = 3.3
                elif mark > 72:
                    grade_point = 3.0
                elif mark > 69:
                    grade_point = 2.7
                elif mark > 66:
                    grade_point = 2.3
                elif mark > 62:
                    grade_point = 2.0
                elif mark > 59:
                    grade_point = 1.7
                elif mark > 56:
                    grade_point = 1.3
                elif mark > 52:
                    grade_point = 1.0
                elif mark > 49:
                    grade_point = 0.7
                else:
                    grade_point = 0

                if course.cworth:
                    total += 1 * grade_point
                    credits += 1
                else:
                    total += .5 * grade_point
                    credits += .5

        if total != 0:
            return total / credits
        return 0.0

    def get_total_credits(self):
        """
        Return total credits
        
        @return: dict of float 
        """
        total = {'Planned': 0, 'Taken': 0}

        courses = self.get_courses()

        for course in courses:
            if course.ctype == 'p':
                total['Planned'] += course.cworth
            elif course.ctype != 'd' and not (course.ctype == 'c' and course.cmark < 50):
                total['Taken'] += course.cworth

        return total

    def _add_pre_courses(self, temp):
        """
        Turns from texted generated values to Course objects
        Only runs at startup

        :param dict temp:
        @return: NoneType
        """

        for key in temp:
            # For each program
            self._course_list[key] = {}
            for item in temp[key]:
                # For each course
                self._course_list[key][item] = Course(temp[key][item][0], lecture_code=temp[key][item][1], ctype=temp[key][item][2], cmark=float(temp[key][item][3]))

                # Tell the tree that the course is taken
                self._update_program_course(self._course_list[key][item], True)

        # Update the nodes in the tree for the new course additions
        self._update_program_nodes()

    def _get_course_info(self, course_code):
        """
        Returns course data from api or None if it either doesn't exist or have no internet

        @param course_code: 
        @return: dict | None

        >>> user = User('Tim', 'Cosby')
        >>> user._get_course_info('CSC148H1F')
        """

        try:
            page = str(PAGE + course_code + '%22&limit=1')
            pr = Request(page)
            pr.add_header('Authorization', 'TVwEIjRZP80vhnY8HhM0OzZCMfydh4lA')

            file = urlopen(pr)
            lines = file.read().decode('utf-8')
            lines = ast.literal_eval(lines)
            print(lines)
            return lines[0]

        except HTTPError:
            # If course does not exist
            return None
        except URLError:
            print('No Internet!')
            return None

    """Program Functions"""

    def add_program(self, program_code):
        """
        Adds a program to the user's profile

        @param str program_code: 
        @return: NoneType
        """

        # Only add if program is not already in
        if program_code in PROGRAM_TREE:
            if program_code not in self._program_list:
                self._program_list.append(program_code)

                # Update database
                WORKSHEET.cell(column=4, row=self._user_row, value=str(self._program_list))
                WORKBOOK.save(FILE_NAME)
        else:
            print('Invalid Program!')

    def remove_program(self, program_code):
        """
        Delete a course from the list

        @param str program_code: Code for the program 
        @return: NoneType
        """

        try:
            self._program_list.remove(program_code)

            # Update database
            WORKSHEET.cell(column=4, row=self._user_row, value=str(self._program_list))
            WORKBOOK.save(FILE_NAME)
        except ValueError:
            print('Program does not exist')

    def get_easiest(self, ratio=True):
        """
        Return easiest to get courses

        @return: list
        """
        totals = {}

        def eval_totals(totals, temp):
            """
            Add to the total need and have

            @param dict of int totals: 
            @param list of CourseNode | CourseNode temp: 
            @return: 
            """

            if temp.requirement is True:
                # If the group requirements have been met
                totals[program][0] += temp.need
                totals[program][1] += temp.need
            else:
                # If the group requirements have yet to be met
                totals[program][0] += temp.need
                totals[program][1] += temp.have

            return totals

        for program in self._user_program_tree:
            # For each program
            totals[program] = [0, 0]  # [Need, Have]

            for course in self._user_program_tree[program]:
                # For each course
                temp = self._user_program_tree[program][course]

                if isinstance(temp, list):
                    # If grouped courses
                    for node in temp:
                        totals.update(eval_totals(totals, node))

                else:
                    # If a CourseNode
                    totals.update(eval_totals(totals, temp))

            fce = self._program_tree[program]['fce']
            if fce[0] is not None:
                # If overall fce needed exists
                temp = fce[0] - totals[program][0] # Get difference between what is needed overall and what is currently needed

                if temp > 0: # If more required fces needed than courses needed to take
                    totals[program][0] += temp

            if fce[1:].count(None) != 4:
                # If a level requirement exists
                fce_temp = [0, 0, 0, 0]  # Levels [1, 2, 3, 4]

                for course in self.get_courses():
                    # For every course taken

                    fce_temp[int(course.ccode[3]) - 1] += course.cworth

                for index in range(4):
                    # For each level requirement
                    if fce[1 + index] is not None:
                        # If the level requirement is not empty
                        temp = fce[1 + index] - fce_temp[index]  # Get the difference

                        if temp > 0:
                            # If needed courses of the level are more than the courses taken of that level
                            totals[program][0] += temp  # Add to the difference to the needed pile

        # Index 0: name, 1: need, 2: have, 3: ratio, 4: need-to-get
        sorted_totals = [[i] + totals[i] for i in totals]
        sorted_totals = [i + [i[2] / i[1]] + [i[1] - i[2]] for i in
                         sorted_totals]

        if ratio:
            index = 3
        else:
            index = 4

        sorted_totals.sort(reverse=True, key=lambda x: x[index])

        return sorted_totals

    def get_program_req(self, program_code):
        """
        Returns a dictionary of the currently needed courses for program <program_code>

        @param str program_code:
        @return: dict

        >>> usr = User('Tim', 'Cosby')
        >>> usr.get_program_req('ASMAJ1423')
        {'BR5': CourseNode('BR5', 1.0, False, None), 'lists': [CourseNode(None, 2, False, {'ENG250Y1': CourseNode('ENG250Y1', 1, False, None), 'HIS271Y1': CourseNode('HIS271Y1', 1, False, None), 'GGR240H1': CourseNode('GGR240H1', 0.5, False, None), 'GGR254H1': CourseNode('GGR254H1', 0.5, False, None), 'POL203Y1': CourseNode('POL203Y1', 1, False, None)}), CourseNode(None, 9, False, {'USA300H1': CourseNode('USA300H1', 0.5, False, None), 'lists': [CourseNode(None, 1, False, {'USA310H1': CourseNode('USA310H1', 0.5, False, None), 'USA311H1': CourseNode('USA311H1', 0.5, False, None), 'USA312H1': CourseNode('USA312H1', 0.5, False, None), 'USA313H1': CourseNode('USA313H1', 0.5, False, None)}), CourseNode(None, 1, False, {'USA400H1': CourseNode('USA400H1', 0.5, False, None), 'USA401H1': CourseNode('USA401H1', 0.5, False, None), 'USA402H1': CourseNode('USA402H1', 0.5, False, None), 'USA403H1': CourseNode('USA403H1', 0.5, False, None)}), CourseNode(None, 1, False, {'CIN490Y1': CourseNode('CIN490Y1', 1, False, None), 'CIN491H1': CourseNode('CIN491H1', 0.5, False, None), 'CIN492H1': CourseNode('CIN492H1', 0.5, False, None)}), CourseNode(None, 1, False, {'ENG434H1': CourseNode('ENG434H1', 0.5, False, None), 'ENG435H1': CourseNode('ENG435H1', 0.5, False, None)})], 'ABS302H1': CourseNode('ABS302H1', 0.5, False, None), 'ABS341H1': CourseNode('ABS341H1', 0.5, False, None), 'ANT365H1': CourseNode('ANT365H1', 0.5, False, None), 'CIN310Y1': CourseNode('CIN310Y1', 1, False, None), 'CIN374Y1': CourseNode('CIN374Y1', 1, False, None), 'CIN334H1': CourseNode('CIN334H1', 0.5, False, None), 'DRM310H1': CourseNode('DRM310H1', 0.5, False, None), 'ENG360H1': CourseNode('ENG360H1', 0.5, False, None), 'ENG363Y1': CourseNode('ENG363Y1', 1, False, None), 'ENG364Y1': CourseNode('ENG364Y1', 1, False, None), 'ENG365H1': CourseNode('ENG365H1', 0.5, False, None), 'ENG368H1': CourseNode('ENG368H1', 0.5, False, None), 'ENG375Y1': CourseNode('ENG375Y1', 1, False, None), 'FAH375H1': CourseNode('FAH375H1', 0.5, False, None), 'GGR336H1': CourseNode('GGR336H1', 0.5, False, None), 'GGR339H1': CourseNode('GGR339H1', 0.5, False, None), 'HIS300H1': CourseNode('HIS300H1', 0.5, False, None), 'HIS310H1': CourseNode('HIS310H1', 0.5, False, None), 'HIS316H1': CourseNode('HIS316H1', 0.5, False, None), 'HIS343Y1': CourseNode('HIS343Y1', 1, False, None), 'HIS365H1': CourseNode('HIS365H1', 0.5, False, None), 'HIS366H1': CourseNode('HIS366H1', 0.5, False, None), 'HIS369H1': CourseNode('HIS369H1', 0.5, False, None), 'HIS370H1': CourseNode('HIS370H1', 0.5, False, None), 'HIS372H1': CourseNode('HIS372H1', 0.5, False, None), 'HIS373H1': CourseNode('HIS373H1', 0.5, False, None), 'HIS374H1': CourseNode('HIS374H1', 0.5, False, None), 'HIS375H1': CourseNode('HIS375H1', 0.5, False, None), 'HIS376H1': CourseNode('HIS376H1', 0.5, False, None), 'HIS377H1': CourseNode('HIS377H1', 0.5, False, None), 'HIS378H1': CourseNode('HIS378H1', 0.5, False, None), 'HIS389H1': CourseNode('HIS389H1', 0.5, False, None), 'HIS393H1': CourseNode('HIS393H1', 0.5, False, None), 'MUS306H1': CourseNode('MUS306H1', 0.5, False, None), 'POL300H1': CourseNode('POL300H1', 0.5, False, None), 'POL326Y1': CourseNode('POL326Y1', 1, False, None), 'RLG315H1': CourseNode('RLG315H1', 0.5, False, None), 'USA494H1': CourseNode('USA494H1', 0.5, False, None), 'USA495Y1': CourseNode('USA495Y1', 1, False, None), 'EAS474H1': CourseNode('EAS474H1', 0.5, False, None), 'ECO423H1': CourseNode('ECO423H1', 0.5, False, None), 'ENG438H1': CourseNode('ENG438H1', 0.5, False, None), 'HIS400H1': CourseNode('HIS400H1', 0.5, False, None), 'HIS401H1': CourseNode('HIS401H1', 0.5, False, None), 'HIS404H1': CourseNode('HIS404H1', 0.5, False, None), 'HIS408Y1': CourseNode('HIS408Y1', 1, False, None), 'HIS436H1': CourseNode('HIS436H1', 0.5, False, None), 'HIS447H1': CourseNode('HIS447H1', 0.5, False, None), 'HIS463H1': CourseNode('HIS463H1', 0.5, False, None), 'HIS473Y1': CourseNode('HIS473Y1', 1, False, None), 'HIS476H1': CourseNode('HIS476H1', 0.5, False, None), 'HIS478H1': CourseNode('HIS478H1', 0.5, False, None), 'HIS479H1': CourseNode('HIS479H1', 0.5, False, None), 'HIS484H1': CourseNode('HIS484H1', 0.5, False, None), 'HIS487H1': CourseNode('HIS487H1', 0.5, False, None), 'POL433H1': CourseNode('POL433H1', 0.5, False, None), 'RLG442H1': CourseNode('RLG442H1', 0.5, False, None), 'WGS435H1': CourseNode('WGS435H1', 0.5, False, None)}), CourseNode(None, 4, False, {'USA300H1': CourseNode('USA300H1', 0.5, False, None), 'lists': [CourseNode(None, 1, False, {'USA310H1': CourseNode('USA310H1', 0.5, False, None), 'USA311H1': CourseNode('USA311H1', 0.5, False, None), 'USA312H1': CourseNode('USA312H1', 0.5, False, None), 'USA313H1': CourseNode('USA313H1', 0.5, False, None)})], 'ABS302H1': CourseNode('ABS302H1', 0.5, False, None), 'ABS341H1': CourseNode('ABS341H1', 0.5, False, None), 'ANT365H1': CourseNode('ANT365H1', 0.5, False, None), 'CIN310Y1': CourseNode('CIN310Y1', 1, False, None), 'CIN374Y1': CourseNode('CIN374Y1', 1, False, None), 'CIN334H1': CourseNode('CIN334H1', 0.5, False, None), 'DRM310H1': CourseNode('DRM310H1', 0.5, False, None), 'ENG360H1': CourseNode('ENG360H1', 0.5, False, None), 'ENG363Y1': CourseNode('ENG363Y1', 1, False, None), 'ENG364Y1': CourseNode('ENG364Y1', 1, False, None), 'ENG365H1': CourseNode('ENG365H1', 0.5, False, None), 'ENG368H1': CourseNode('ENG368H1', 0.5, False, None), 'ENG375Y1': CourseNode('ENG375Y1', 1, False, None), 'FAH375H1': CourseNode('FAH375H1', 0.5, False, None), 'GGR336H1': CourseNode('GGR336H1', 0.5, False, None), 'GGR339H1': CourseNode('GGR339H1', 0.5, False, None), 'HIS300H1': CourseNode('HIS300H1', 0.5, False, None), 'HIS310H1': CourseNode('HIS310H1', 0.5, False, None), 'HIS316H1': CourseNode('HIS316H1', 0.5, False, None), 'HIS343Y1': CourseNode('HIS343Y1', 1, False, None), 'HIS365H1': CourseNode('HIS365H1', 0.5, False, None), 'HIS366H1': CourseNode('HIS366H1', 0.5, False, None), 'HIS369H1': CourseNode('HIS369H1', 0.5, False, None), 'HIS370H1': CourseNode('HIS370H1', 0.5, False, None), 'HIS372H1': CourseNode('HIS372H1', 0.5, False, None), 'HIS373H1': CourseNode('HIS373H1', 0.5, False, None), 'HIS374H1': CourseNode('HIS374H1', 0.5, False, None), 'HIS375H1': CourseNode('HIS375H1', 0.5, False, None), 'HIS376H1': CourseNode('HIS376H1', 0.5, False, None), 'HIS377H1': CourseNode('HIS377H1', 0.5, False, None), 'HIS378H1': CourseNode('HIS378H1', 0.5, False, None), 'HIS389H1': CourseNode('HIS389H1', 0.5, False, None), 'HIS393H1': CourseNode('HIS393H1', 0.5, False, None), 'MUS306H1': CourseNode('MUS306H1', 0.5, False, None), 'POL300H1': CourseNode('POL300H1', 0.5, False, None), 'POL326Y1': CourseNode('POL326Y1', 1, False, None), 'RLG315H1': CourseNode('RLG315H1', 0.5, False, None)}), CourseNode(None, 13, False, {'USA200H1': CourseNode('USA200H1', 0.5, False, None), 'USA300H1': CourseNode('USA300H1', 0.5, False, None), 'lists': [CourseNode(None, 1, False, {'USA310H1': CourseNode('USA310H1', 0.5, False, None), 'USA311H1': CourseNode('USA311H1', 0.5, False, None), 'USA312H1': CourseNode('USA312H1', 0.5, False, None), 'USA313H1': CourseNode('USA313H1', 0.5, False, None)}), CourseNode(None, 1, False, {'USA400H1': CourseNode('USA400H1', 0.5, False, None), 'USA401H1': CourseNode('USA401H1', 0.5, False, None), 'USA402H1': CourseNode('USA402H1', 0.5, False, None), 'USA403H1': CourseNode('USA403H1', 0.5, False, None)}), CourseNode(None, 1, False, {'CIN490Y1': CourseNode('CIN490Y1', 1, False, None), 'CIN491H1': CourseNode('CIN491H1', 0.5, False, None), 'CIN492H1': CourseNode('CIN492H1', 0.5, False, None)}), CourseNode(None, 1, False, {'ENG434H1': CourseNode('ENG434H1', 0.5, False, None), 'ENG435H1': CourseNode('ENG435H1', 0.5, False, None)})], 'USA494H1': CourseNode('USA494H1', 0.5, False, None), 'USA495Y1': CourseNode('USA495Y1', 1, False, None), 'ABS302H1': CourseNode('ABS302H1', 0.5, False, None), 'ABS341H1': CourseNode('ABS341H1', 0.5, False, None), 'ANT365H1': CourseNode('ANT365H1', 0.5, False, None), 'CIN270Y1': CourseNode('CIN270Y1', 1, False, None), 'CIN211H1': CourseNode('CIN211H1', 0.5, False, None), 'CIN230H1': CourseNode('CIN230H1', 0.5, False, None), 'CIN310Y1': CourseNode('CIN310Y1', 1, False, None), 'CIN374Y1': CourseNode('CIN374Y1', 1, False, None), 'CIN334H1': CourseNode('CIN334H1', 0.5, False, None), 'DRM310H1': CourseNode('DRM310H1', 0.5, False, None), 'EAS474H1': CourseNode('EAS474H1', 0.5, False, None), 'ECO423H1': CourseNode('ECO423H1', 0.5, False, None), 'ENG250Y1': CourseNode('ENG250Y1', 1, False, None), 'ENG254Y1': CourseNode('ENG254Y1', 1, False, None), 'ENG360H1': CourseNode('ENG360H1', 0.5, False, None), 'ENG363Y1': CourseNode('ENG363Y1', 1, False, None), 'ENG364Y1': CourseNode('ENG364Y1', 1, False, None), 'ENG365H1': CourseNode('ENG365H1', 0.5, False, None), 'ENG368H1': CourseNode('ENG368H1', 0.5, False, None), 'ENG375Y1': CourseNode('ENG375Y1', 1, False, None), 'ENG438H1': CourseNode('ENG438H1', 0.5, False, None), 'FAH375H1': CourseNode('FAH375H1', 0.5, False, None), 'GGR240H1': CourseNode('GGR240H1', 0.5, False, None), 'GGR254H1': CourseNode('GGR254H1', 0.5, False, None), 'GGR336H1': CourseNode('GGR336H1', 0.5, False, None), 'GGR339H1': CourseNode('GGR339H1', 0.5, False, None), 'HIS106Y1': CourseNode('HIS106Y1', 1, False, None), 'HIS202H1': CourseNode('HIS202H1', 0.5, False, None), 'HIS271Y1': CourseNode('HIS271Y1', 1, False, None), 'HIS300H1': CourseNode('HIS300H1', 0.5, False, None), 'HIS310H1': CourseNode('HIS310H1', 0.5, False, None), 'HIS316H1': CourseNode('HIS316H1', 0.5, False, None), 'HIS343Y1': CourseNode('HIS343Y1', 1, False, None), 'HIS365H1': CourseNode('HIS365H1', 0.5, False, None), 'HIS366H1': CourseNode('HIS366H1', 0.5, False, None), 'HIS369H1': CourseNode('HIS369H1', 0.5, False, None), 'HIS370H1': CourseNode('HIS370H1', 0.5, False, None), 'HIS372H1': CourseNode('HIS372H1', 0.5, False, None), 'HIS373H1': CourseNode('HIS373H1', 0.5, False, None), 'HIS374H1': CourseNode('HIS374H1', 0.5, False, None), 'HIS375H1': CourseNode('HIS375H1', 0.5, False, None), 'HIS376H1': CourseNode('HIS376H1', 0.5, False, None), 'HIS377H1': CourseNode('HIS377H1', 0.5, False, None), 'HIS378H1': CourseNode('HIS378H1', 0.5, False, None), 'HIS389H1': CourseNode('HIS389H1', 0.5, False, None), 'HIS393H1': CourseNode('HIS393H1', 0.5, False, None), 'HIS400H1': CourseNode('HIS400H1', 0.5, False, None), 'HIS401H1': CourseNode('HIS401H1', 0.5, False, None), 'HIS404H1': CourseNode('HIS404H1', 0.5, False, None), 'HIS408Y1': CourseNode('HIS408Y1', 1, False, None), 'HIS436H1': CourseNode('HIS436H1', 0.5, False, None), 'HIS447H1': CourseNode('HIS447H1', 0.5, False, None), 'HIS463H1': CourseNode('HIS463H1', 0.5, False, None), 'HIS473Y1': CourseNode('HIS473Y1', 1, False, None), 'HIS476H1': CourseNode('HIS476H1', 0.5, False, None), 'HIS478H1': CourseNode('HIS478H1', 0.5, False, None), 'HIS479H1': CourseNode('HIS479H1', 0.5, False, None), 'HIS484H1': CourseNode('HIS484H1', 0.5, False, None), 'HIS487H1': CourseNode('HIS487H1', 0.5, False, None), 'MUS306H1': CourseNode('MUS306H1', 0.5, False, None), 'POL203Y1': CourseNode('POL203Y1', 1, False, None), 'POL300H1': CourseNode('POL300H1', 0.5, False, None), 'POL326Y1': CourseNode('POL326Y1', 1, False, None), 'POL433H1': CourseNode('POL433H1', 0.5, False, None), 'RLG315H1': CourseNode('RLG315H1', 0.5, False, None), 'RLG442H1': CourseNode('RLG442H1', 0.5, False, None), 'WGS435H1': CourseNode('WGS435H1', 0.5, False, None), 'VIC132H1': CourseNode('VIC132H1', 0.5, False, None), 'VIC130H1': CourseNode('VIC130H1', 0.5, False, None)}), CourseNode(None, 4, False, {'lists': [CourseNode(None, 2, False, {'USA200H1': CourseNode('USA200H1', 0.5, False, None), 'USA300H1': CourseNode('USA300H1', 0.5, False, None), 'lists': [CourseNode(None, 1, False, {'USA310H1': CourseNode('USA310H1', 0.5, False, None), 'USA311H1': CourseNode('USA311H1', 0.5, False, None), 'USA312H1': CourseNode('USA312H1', 0.5, False, None), 'USA313H1': CourseNode('USA313H1', 0.5, False, None)}), CourseNode(None, 1, False, {'USA400H1': CourseNode('USA400H1', 0.5, False, None), 'USA401H1': CourseNode('USA401H1', 0.5, False, None), 'USA402H1': CourseNode('USA402H1', 0.5, False, None), 'USA403H1': CourseNode('USA403H1', 0.5, False, None)})], 'USA494H1': CourseNode('USA494H1', 0.5, False, None), 'USA495Y1': CourseNode('USA495Y1', 1, False, None)}), CourseNode(None, 2, False, {'ABS302H1': CourseNode('ABS302H1', 0.5, False, None), 'ABS341H1': CourseNode('ABS341H1', 0.5, False, None)}), CourseNode(None, 2, False, {'ANT365H1': CourseNode('ANT365H1', 0.5, False, None)}), CourseNode(None, 2, False, {'CIN270Y1': CourseNode('CIN270Y1', 1, False, None), 'CIN211H1': CourseNode('CIN211H1', 0.5, False, None), 'CIN230H1': CourseNode('CIN230H1', 0.5, False, None), 'CIN310Y1': CourseNode('CIN310Y1', 1, False, None), 'CIN374Y1': CourseNode('CIN374Y1', 1, False, None), 'CIN334H1': CourseNode('CIN334H1', 0.5, False, None), 'lists': [CourseNode(None, 1, False, {'CIN490Y1': CourseNode('CIN490Y1', 1, False, None), 'CIN491H1': CourseNode('CIN491H1', 0.5, False, None), 'CIN492H1': CourseNode('CIN492H1', 0.5, False, None)})]}), CourseNode(None, 2, False, {'DRM310H1': CourseNode('DRM310H1', 0.5, False, None)}), CourseNode(None, 2, False, {'EAS474H1': CourseNode('EAS474H1', 0.5, False, None)}), CourseNode(None, 2, False, {'ECO423H1': CourseNode('ECO423H1', 0.5, False, None)}), CourseNode(None, 2, False, {'ENG250Y1': CourseNode('ENG250Y1', 1, False, None), 'ENG254Y1': CourseNode('ENG254Y1', 1, False, None), 'ENG360H1': CourseNode('ENG360H1', 0.5, False, None), 'ENG363Y1': CourseNode('ENG363Y1', 1, False, None), 'ENG364Y1': CourseNode('ENG364Y1', 1, False, None), 'ENG365H1': CourseNode('ENG365H1', 0.5, False, None), 'ENG368H1': CourseNode('ENG368H1', 0.5, False, None), 'ENG375Y1': CourseNode('ENG375Y1', 1, False, None), 'lists': [CourseNode(None, 1, False, {'ENG434H1': CourseNode('ENG434H1', 0.5, False, None), 'ENG435H1': CourseNode('ENG435H1', 0.5, False, None)})], 'ENG438H1': CourseNode('ENG438H1', 0.5, False, None)}), CourseNode(None, 2, False, {'FAH375H1': CourseNode('FAH375H1', 0.5, False, None)}), CourseNode(None, 2, False, {'GGR240H1': CourseNode('GGR240H1', 0.5, False, None), 'GGR254H1': CourseNode('GGR254H1', 0.5, False, None), 'GGR336H1': CourseNode('GGR336H1', 0.5, False, None), 'GGR339H1': CourseNode('GGR339H1', 0.5, False, None)}), CourseNode(None, 2, False, {'HIS106Y1': CourseNode('HIS106Y1', 1, False, None), 'HIS202H1': CourseNode('HIS202H1', 0.5, False, None), 'HIS271Y1': CourseNode('HIS271Y1', 1, False, None), 'HIS300H1': CourseNode('HIS300H1', 0.5, False, None), 'HIS310H1': CourseNode('HIS310H1', 0.5, False, None), 'HIS316H1': CourseNode('HIS316H1', 0.5, False, None), 'HIS343Y1': CourseNode('HIS343Y1', 1, False, None), 'HIS365H1': CourseNode('HIS365H1', 0.5, False, None), 'HIS366H1': CourseNode('HIS366H1', 0.5, False, None), 'HIS369H1': CourseNode('HIS369H1', 0.5, False, None), 'HIS370H1': CourseNode('HIS370H1', 0.5, False, None), 'HIS372H1': CourseNode('HIS372H1', 0.5, False, None), 'HIS373H1': CourseNode('HIS373H1', 0.5, False, None), 'HIS374H1': CourseNode('HIS374H1', 0.5, False, None), 'HIS375H1': CourseNode('HIS375H1', 0.5, False, None), 'HIS376H1': CourseNode('HIS376H1', 0.5, False, None), 'HIS377H1': CourseNode('HIS377H1', 0.5, False, None), 'HIS378H1': CourseNode('HIS378H1', 0.5, False, None), 'HIS389H1': CourseNode('HIS389H1', 0.5, False, None), 'HIS393H1': CourseNode('HIS393H1', 0.5, False, None), 'HIS400H1': CourseNode('HIS400H1', 0.5, False, None), 'HIS401H1': CourseNode('HIS401H1', 0.5, False, None), 'HIS404H1': CourseNode('HIS404H1', 0.5, False, None), 'HIS408Y1': CourseNode('HIS408Y1', 1, False, None), 'HIS436H1': CourseNode('HIS436H1', 0.5, False, None), 'HIS447H1': CourseNode('HIS447H1', 0.5, False, None), 'HIS463H1': CourseNode('HIS463H1', 0.5, False, None), 'HIS473Y1': CourseNode('HIS473Y1', 1, False, None), 'HIS476H1': CourseNode('HIS476H1', 0.5, False, None), 'HIS478H1': CourseNode('HIS478H1', 0.5, False, None), 'HIS479H1': CourseNode('HIS479H1', 0.5, False, None), 'HIS484H1': CourseNode('HIS484H1', 0.5, False, None), 'HIS487H1': CourseNode('HIS487H1', 0.5, False, None)}), CourseNode(None, 2, False, {'HIS106Y1': CourseNode('HIS106Y1', 1, False, None), 'HIS202H1': CourseNode('HIS202H1', 0.5, False, None), 'HIS271Y1': CourseNode('HIS271Y1', 1, False, None), 'HIS300H1': CourseNode('HIS300H1', 0.5, False, None), 'HIS310H1': CourseNode('HIS310H1', 0.5, False, None), 'HIS316H1': CourseNode('HIS316H1', 0.5, False, None), 'HIS343Y1': CourseNode('HIS343Y1', 1, False, None), 'HIS365H1': CourseNode('HIS365H1', 0.5, False, None), 'HIS366H1': CourseNode('HIS366H1', 0.5, False, None), 'HIS369H1': CourseNode('HIS369H1', 0.5, False, None), 'HIS370H1': CourseNode('HIS370H1', 0.5, False, None), 'HIS372H1': CourseNode('HIS372H1', 0.5, False, None), 'HIS373H1': CourseNode('HIS373H1', 0.5, False, None), 'HIS374H1': CourseNode('HIS374H1', 0.5, False, None), 'HIS375H1': CourseNode('HIS375H1', 0.5, False, None), 'HIS376H1': CourseNode('HIS376H1', 0.5, False, None), 'HIS377H1': CourseNode('HIS377H1', 0.5, False, None), 'HIS378H1': CourseNode('HIS378H1', 0.5, False, None), 'HIS389H1': CourseNode('HIS389H1', 0.5, False, None), 'HIS393H1': CourseNode('HIS393H1', 0.5, False, None), 'HIS400H1': CourseNode('HIS400H1', 0.5, False, None), 'HIS401H1': CourseNode('HIS401H1', 0.5, False, None), 'HIS404H1': CourseNode('HIS404H1', 0.5, False, None), 'HIS408Y1': CourseNode('HIS408Y1', 1, False, None), 'HIS436H1': CourseNode('HIS436H1', 0.5, False, None), 'HIS447H1': CourseNode('HIS447H1', 0.5, False, None), 'HIS463H1': CourseNode('HIS463H1', 0.5, False, None), 'HIS473Y1': CourseNode('HIS473Y1', 1, False, None), 'HIS476H1': CourseNode('HIS476H1', 0.5, False, None), 'HIS478H1': CourseNode('HIS478H1', 0.5, False, None), 'HIS479H1': CourseNode('HIS479H1', 0.5, False, None), 'HIS484H1': CourseNode('HIS484H1', 0.5, False, None), 'HIS487H1': CourseNode('HIS487H1', 0.5, False, None)}), CourseNode(None, 2, False, {'MUS306H1': CourseNode('MUS306H1', 0.5, False, None)}), CourseNode(None, 2, False, {'POL203Y1': CourseNode('POL203Y1', 1, False, None), 'POL300H1': CourseNode('POL300H1', 0.5, False, None), 'POL326Y1': CourseNode('POL326Y1', 1, False, None), 'POL433H1': CourseNode('POL433H1', 0.5, False, None)}), CourseNode(None, 2, False, {'RLG315H1': CourseNode('RLG315H1', 0.5, False, None), 'RLG442H1': CourseNode('RLG442H1', 0.5, False, None)}), CourseNode(None, 2, False, {'WGS435H1': CourseNode('WGS435H1', 0.5, False, None)}), CourseNode(None, 2, False, {'VIC132H1': CourseNode('VIC132H1', 0.5, False, None), 'VIC130H1': CourseNode('VIC130H1', 0.5, False, None)})]}), CourseNode(None, 1, False, {'lists': [CourseNode(None, 1, False, {'USA400H1': CourseNode('USA400H1', 0.5, False, None), 'USA401H1': CourseNode('USA401H1', 0.5, False, None), 'USA402H1': CourseNode('USA402H1', 0.5, False, None), 'USA403H1': CourseNode('USA403H1', 0.5, False, None)}), CourseNode(None, 1, False, {'CIN490Y1': CourseNode('CIN490Y1', 1, False, None), 'CIN491H1': CourseNode('CIN491H1', 0.5, False, None), 'CIN492H1': CourseNode('CIN492H1', 0.5, False, None)}), CourseNode(None, 1, False, {'ENG434H1': CourseNode('ENG434H1', 0.5, False, None), 'ENG435H1': CourseNode('ENG435H1', 0.5, False, None)})], 'USA494H1': CourseNode('USA494H1', 0.5, False, None), 'USA495Y1': CourseNode('USA495Y1', 1, False, None), 'EAS474H1': CourseNode('EAS474H1', 0.5, False, None), 'ECO423H1': CourseNode('ECO423H1', 0.5, False, None), 'ENG438H1': CourseNode('ENG438H1', 0.5, False, None), 'HIS400H1': CourseNode('HIS400H1', 0.5, False, None), 'HIS401H1': CourseNode('HIS401H1', 0.5, False, None), 'HIS404H1': CourseNode('HIS404H1', 0.5, False, None), 'HIS408Y1': CourseNode('HIS408Y1', 1, False, None), 'HIS436H1': CourseNode('HIS436H1', 0.5, False, None), 'HIS447H1': CourseNode('HIS447H1', 0.5, False, None), 'HIS463H1': CourseNode('HIS463H1', 0.5, False, None), 'HIS473Y1': CourseNode('HIS473Y1', 1, False, None), 'HIS476H1': CourseNode('HIS476H1', 0.5, False, None), 'HIS478H1': CourseNode('HIS478H1', 0.5, False, None), 'HIS479H1': CourseNode('HIS479H1', 0.5, False, None), 'HIS484H1': CourseNode('HIS484H1', 0.5, False, None), 'HIS487H1': CourseNode('HIS487H1', 0.5, False, None), 'POL433H1': CourseNode('POL433H1', 0.5, False, None), 'RLG442H1': CourseNode('RLG442H1', 0.5, False, None), 'WGS435H1': CourseNode('WGS435H1', 0.5, False, None)})], 'USA300H1': CourseNode('USA300H1', 0.5, False, None)}
        """
        return self._user_program_tree[program_code]

    def get_programs(self):
        """
        Return a list of taken programs

        >>> usr = User('Tim', 'Cosby')
        >>> usr.add_program('ASMAJ0135')
        >>> usr.get_programs()
        ['ASMAJ0135']
        """
        return self._program_list

    def get_program_name(self, program_code):
        return self._program_tree[program_code]['name']

    def get_program_type(self, program_code):
        return self._program_tree[program_code]['type']

    def get_program_description(self, program_code):
        return self._program_tree[program_code]['description']

    def _update_inner_nodes(self, root=None):
        """
        Update node if not a leaf and has met the requirements of the node
        Note: Only used by _update_program_nodes

        @param dict root: 
        @return: NoneType
        """

        for child in root:
            # For CourseNode in dict
            if child.course_code is None:
                # If child is a node, get the total credits from its child courses
                child.have = self._update_program_nodes(root=child.children)

                if child.have >= child.need:
                    # If the sum credits from its child courses satisfies its requirements
                    child.requirement = True

    def _update_program_nodes(self, root=None):
        """
        Update nodes in the program tree to accommodate for changes
        to courses
        Note: Nodes are only groups of courses
              Single courses are leafs

        @param None | dict root: root dictionary looking 
        @return: int

        {<Program Name>: {'lists': [CourseNode(None, 2, False, {'CSC108H1': CourseNode('CSC108H1', 1, False, None), 'CSC148H1': CourseNode('CSC148H1', 1, False, None)}]}}
        *Takes CSC108*
        {<Program Name>: {'lists': [CourseNode(None, 2, False, {'CSC108H1': CourseNode('CSC108H1', 1, True, None), 'CSC148H1': CourseNode('CSC148H1', 1, False, None)}]}}
        *Takes CSC148*
        {<Program Name>: {'lists': [CourseNode(None, 2, True, {'CSC108H1': CourseNode('CSC108H1', 1, True, None), 'CSC148H1': CourseNode('CSC148H1', 1, True, None)}]}}
        Note: Node gets updated to True as the internal courses met its requirements of 2 credits
        """
        total = 0

        if root is None:
            # If first iteration
            for program in self._user_program_tree:
                # Recurse through every program
                self._update_program_nodes(root=self._user_program_tree[program])

        else:
            # If a single program or node dict
            try:
                # Tries to recurse through the groups of courses
                self._update_inner_nodes(root=root['lists'])

                # Adds up the total credits taken in the group
                total += sum(
                    [i.need for i in root['lists'] if i.requirement is True])

            except KeyError:
                pass

            # Adds up the total single credits taken in the dictionary
            total += sum([root[i].need for i in root if
                          i != 'list' and isinstance(root[i], CourseNode) and
                          root[i].requirement is True])

            return total

    def _update_program_course(self, course, add=True):
        """
        Updates the tree with new requirements
        aka add or remove a course taken value from tree

        Note: Only adds courses if they have not been failed or dropped

        @param Course course: Course's object
        @param bool add: True if adding False if removing 
        @return: NoneType
        """

        # If there is a **** course in cache, search up if the inputted course also meets the requirements (Would need to just update its have category if it doesn't)

        try:
            cached_couse = self._program_course_cache[course.ccode[:-1]]

            if course.ctype == 'd' or (course.ctype == 'c' and course.cmark < 50) or not add:
                # If dropped or have a failing grade or removing
                cached_couse.requirement = False
            else:
                # Completes requirement
                cached_couse.requirement = True
        except KeyError:
            pass

        self._update_breadths()

    def _update_breadths(self):
        """
        Update breadth requirements in the tree
        Ex: BR5 after adding a course with Breadth 5
        
        @return: NoneType
        """

        breadths = self.get_breadths()

        for index in range(1, 6):
            try:
                breadth_node = self._program_course_cache['BR' + str(index)]

                breadth_node.have = breadths['Taken'][index-1] + breadths['Planned'][index-1]

                if breadth_node.have >= breadth_node.need:
                    breadth_node.requirement = True
                else:
                    breadth_node.requirement = False
            except KeyError:
                pass



if __name__ == '__main__':
    details = None

    while True:
        # Run till correct login details have been entered

        details = User(input('Enter Username: '), input('Enter Password: '))

        if not details.get_loggedin():
            del details  # Delete the failed object
        else:
            break

    while True:
        #print(details.get_program_req('ASMAJ0135'))
        print([i.ccode for i in details.get_courses()])
        print(details.get_easiest())
        course_code = input('Enter a course: ')
        ctype = input('Type: ')
        cmark = input('Mark: ')
        lecture_code = input('Lecture: ')

        if course_code != '' and ctype != '' and cmark != '' and lecture_code != '':
            details.add_course(course_code, lecture_code=lecture_code, ctype=ctype, cmark=float(cmark))
